import csv
import openpyxl as xl
from pathlib import Path
import PySimpleGUI as sg

# PySimpleGUI layout
layout = [
    [sg.Text("Select CSV file for processing:")],
    [sg.InputText(key="CSV_FILE"), sg.FileBrowse()],
    [sg.Text("Select Excel output file:")],
    [sg.InputText(key="EXCEL_FILE"), sg.FileBrowse()],
    [sg.Button("Process")],
]

window = sg.Window("Dues Processing Tool", layout)

def process_dues():
    order_file_path = values["CSV_FILE"]
    dues_book_path = values["EXCEL_FILE"]

    # Your existing code for processing CSV and Excel goes here
    dues_book_exists = True if Path(dues_book_path).is_file() else False
    dues_book = xl.load_workbook(dues_book_path) if dues_book_exists else xl.Workbook()

    dues_book_exists = True if Path(dues_book_path).is_file() else False
    dues_book = xl.load_workbook(dues_book_path) if dues_book_exists else xl.Workbook()
    sheet_lookup = {
        "Neither current staff or student": "Alumni",
        "Current GT grad student": "Grad",
        "Current GT undergrad student": "Undergrad",
        "Current GT faculty/staff": "Employees"
    }
    new_orders = []

    if not dues_book_exists:
        dues_book.create_sheet("Alumni")
        dues_book.create_sheet("Grad")
        dues_book.create_sheet("Undergrad")
        dues_book.create_sheet("Employees")
        dues_book.remove(dues_book["Sheet"])

        for sheet in dues_book.worksheets:
            sheet.cell(1, 1, "Name")
            sheet.cell(1, 5, "Email")
        dues_book.save(dues_book_path)

    with open(order_file_path, "r", encoding="utf-8") as order_file:
        order_dict = csv.DictReader(order_file)
        for order in order_dict:
            if (order["Fulfillment Status"] == "New" and order["Item Name"] == "Membership Dues"):
                new_orders.append(order)

    for order in new_orders:
        exists = False
        sheet_name = sheet_lookup[order["Item Variation"]]
        sheet = dues_book[sheet_name]
        for cell in sheet["A"]:
            if (cell.value == order["Recipient Name"]):
                exists = True
        sheet.append([order["Recipient Name"], None, None, None, order["Recipient Email"]]) if not exists else None

    dues_book.save(dues_book_path)

while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED:
        break
    elif event == "Process":
        try:
            process_dues()
            sg.popup("Processing complete!", "New data has been added to the Excel file.")
        except:
            sg.popup("Processing Error!")

window.close()