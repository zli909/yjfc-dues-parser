import csv
import openpyxl as xl
from pathlib import Path
import PySimpleGUI as sg

# PySimpleGUI layout
layout = [
    [sg.Text("Select CSV file for processing:")],
    [sg.InputText(key="CSV_FILE"), sg.FileBrowse(file_types=(("CSV Files", "*.csv"), ("All Files", "*.*")))],
    [sg.Text("Select Excel output file:")],
    [sg.InputText(key="EXCEL_FILE"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))],
    [sg.Button("Process")],
]

window = sg.Window("Dues Processing Tool", layout)

sheets = ["Alumni", "Grad", "Undergrad", "Employees"]

sheet_lookup = {
    "Neither Current Staff or Student": "Alumni",
    "Current GT Grad Student": "Grad",
    "Current GT Undergrad Student": "Undergrad",
    "Current GT Faculty/Staff": "Employees"
}

def find_member_type(item_varation):
    for key in sheet_lookup:
        if key in item_varation:
            return key

def process_dues():
    order_file_path = values["CSV_FILE"]
    dues_book_path = values["EXCEL_FILE"]

    # Your existing code for processing CSV and Excel goes here
    dues_book = xl.load_workbook(dues_book_path)
    new_orders = []

    for sheetname in dues_book.sheetnames:
        if sheetname not in sheets:
            dues_book.remove(dues_book[sheetname])

    new_sheets = [sheet for sheet in sheets if sheet not in dues_book.sheetnames]

    for sheet in new_sheets:
        dues_book.create_sheet(sheet)
        dues_book[sheet].cell(1, 1, "Name (Semester)")
        dues_book[sheet].cell(1, 2, "Name (Annual)")
        dues_book[sheet].cell(1, 5, "Email (Semester)")
        dues_book[sheet].cell(1, 6, "Email (Annual)")

    dues_book.save(dues_book_path)

    with open(order_file_path, "r", encoding="utf-8") as order_file:
        order_dict = csv.DictReader(order_file)
        for order in order_dict:
            if (order["Fulfillment Status"] == "New" and order["Item Name"] == "Membership Dues"):
                new_orders.append(order)

    for order in new_orders:
        exists = False
        member_type = find_member_type(order["Item Variation"])
        sheet_name = sheet_lookup[member_type]
        sheet = dues_book[sheet_name]
        name_column_for_dues_type = None
        email_column_for_dues_type = None
        next_empty_name_cell = None
        next_empty_email_cell = None

        # Search the semesterly dues members
        for cell in sheet["A"]:
            if (cell.value == order["Recipient Name"]):
                exists = True
        # Search the annual dues members
        for cell in sheet["B"]:
            if (cell.value == order["Recipient Name"]):
                exists = True

        if not exists:
            if "Annual" in order["Item Variation"]:
                name_column_for_dues_type = 2
                email_column_for_dues_type = 6
            else:
                name_column_for_dues_type = 1
                email_column_for_dues_type = 5
            
            # find the next cell to put the member name in
            for row in range(1, sheet.max_row + 2):
                cell = sheet.cell(row=row, column=name_column_for_dues_type)
                if cell.value is None:
                    next_empty_name_cell = cell
                    break
            # find the next cell to put the member email in
            for row in range(1, sheet.max_row + 2):
                cell = sheet.cell(row=row, column=email_column_for_dues_type)
                if cell.value is None:
                    next_empty_email_cell = cell
                    break
            
            next_empty_name_cell.value = order["Recipient Name"]
            next_empty_email_cell.value = order["Recipient Email"]

    dues_book.save(dues_book_path)

while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED:
        break
    elif event == "Process":
        try:
            process_dues()
            sg.popup("Processing complete!", "New data has been added to the Excel file.")
        except Exception as e:
            sg.popup("Processing Error! " + str(e))

window.close()